"""Excel exports via OpenPyXL. Fully real, no external credentials needed."""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _cell_value(value):
    # openpyxl raises on timezone-aware datetimes ("Excel does not support timezones
    # in datetimes") — every timestamp column in this app is tz-aware (DateTime(timezone=True)),
    # so any row passing one through unchanged crashes the export. Strip the tzinfo
    # (Excel has no timezone concept anyway) rather than requiring every call site to
    # remember to do this themselves.
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _write_sheet(ws, columns: list[str], rows: list[list]) -> None:
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_cell_value(value))

    # Auto-size each column to its content (capped so one long outlier value can't
    # blow out the whole sheet), and freeze the header row so it stays visible while
    # scrolling through a large export.
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(str(col_name))
        for row in rows:
            if col_idx - 1 < len(row) and row[col_idx - 1] is not None:
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
    ws.freeze_panes = "A2"

    # Column filter/sort dropdowns on the header row, for government-report-style
    # spreadsheets where reviewers commonly filter/sort by column (e.g. status,
    # region) rather than reading the whole export top to bottom.
    if rows:
        ws.auto_filter.ref = ws.dimensions


def build_excel_report(title: str, columns: list[str], rows: list[list]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Report"
    _write_sheet(ws, columns, rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_multi_sheet_excel_report(sheets: list[tuple[str, list[str], list[list]]]) -> io.BytesIO:
    """sheets: list of (sheet_title, columns, rows) tuples — one sheet per tuple.
    Used where multiple related tables need to land in a single workbook (e.g. a
    dashboard export with a summary sheet plus one sheet per analytics series)."""
    wb = Workbook()
    wb.remove(wb.active)
    for title, columns, rows in sheets:
        ws = wb.create_sheet(title=(title[:31] or "Sheet"))
        _write_sheet(ws, columns, rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
